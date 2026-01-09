# db/database.py
import os
import sqlite3
import psycopg2
from psycopg2.extras import RealDictCursor
from datetime import datetime, timedelta

# Получаем URL базы из переменных окружения (Railway задаёт автоматически)
DATABASE_URL = os.getenv("DATABASE_URL")

def get_db_connection():
    """Создаёт подключение к базе данных"""
    if DATABASE_URL:
        # Используем PostgreSQL
        return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    else:
        # Локально — SQLite
        conn = sqlite3.connect("finance.db")
        conn.row_factory = sqlite3.Row
        return conn

def init_db():
    """Создаёт таблицы в зависимости от типа базы"""
    conn = get_db_connection()
    cur = conn.cursor()

    if DATABASE_URL:
        # PostgreSQL
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                tg_user_id BIGINT UNIQUE,
                created_at TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS transactions (
                id SERIAL PRIMARY,
                user_id INTEGER REFERENCES users(id),
                type VARCHAR(10),
                amount NUMERIC,
                category VARCHAR(50),
                description TEXT,
                created_at TIMESTAMP
            )
        """)
    else:
        # SQLite
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tg_user_id INTEGER UNIQUE,
                created_at TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                type TEXT,
                amount REAL,
                category TEXT,
                description TEXT,
                created_at TEXT,
                FOREIGN KEY(user_id) REFERENCES users(id)
            )
        """)

    conn.commit()
    conn.close()

def add_user(tg_user_id: int):
    conn = get_db_connection()
    cur = conn.cursor()
    if DATABASE_URL:
        cur.execute(
            "INSERT INTO users (tg_user_id, created_at) VALUES (%s, %s) ON CONFLICT (tg_user_id) DO NOTHING",
            (tg_user_id, datetime.utcnow().isoformat())
        )
    else:
        cur.execute(
            "INSERT OR IGNORE INTO users (tg_user_id, created_at) VALUES (?, ?)",
            (tg_user_id, datetime.utcnow().isoformat())
        )
    conn.commit()
    conn.close()

def get_user_id(tg_user_id: int):
    conn = get_db_connection()
    cur = conn.cursor()
    if DATABASE_URL:
        cur.execute("SELECT id FROM users WHERE tg_user_id = %s", (tg_user_id,))
    else:
        cur.execute("SELECT id FROM users WHERE tg_user_id = ?", (tg_user_id,))
    row = cur.fetchone()
    conn.close()
    return row["id"] if row else None if DATABASE_URL else (row[0] if row else None)

def add_transaction(tg_user_id: int, trans_type: str, amount: float, category: str, description: str = ""):
    user_id = get_user_id(tg_user_id)
    if user_id is None:
        add_user(tg_user_id)
        user_id = get_user_id(tg_user_id)

    conn = get_db_connection()
    cur = conn.cursor()
    if DATABASE_URL:
        cur.execute("""
            INSERT INTO transactions (user_id, type, amount, category, description, created_at)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (user_id, trans_type, amount, category, description, datetime.utcnow().isoformat()))
    else:
        cur.execute("""
            INSERT INTO transactions (user_id, type, amount, category, description, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (user_id, trans_type, amount, category, description, datetime.utcnow().isoformat()))
    conn.commit()
    conn.close()

# === ФУНКЦИИ ОТЧЁТОВ ===

def _execute_query(query, params):
    """Универсальный метод выполнения запроса"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(query, params)
    row = cur.fetchone()
    conn.close()
    return row

def get_daily_summary(user_id: int):
    today = datetime.utcnow().date().isoformat()
    if DATABASE_URL:
        row = _execute_query("""
            SELECT 
                COALESCE(SUM(CASE WHEN type = 'income' THEN amount ELSE 0 END), 0) AS income,
                COALESCE(SUM(CASE WHEN type = 'expense' THEN amount ELSE 0 END), 0) AS expense
            FROM transactions
            WHERE user_id = %s AND created_at::date = %s
        """, (user_id, today))
    else:
        row = _execute_query("""
            SELECT 
                SUM(CASE WHEN type = 'income' THEN amount ELSE 0 END) AS income,
                SUM(CASE WHEN type = 'expense' THEN amount ELSE 0 END) AS expense
            FROM transactions
            WHERE user_id = ? AND DATE(created_at) = ?
        """, (user_id, today))
    
    income = (row["income"] if DATABASE_URL else row[0]) or 0
    expense = (row["expense"] if DATABASE_URL else row[1]) or 0
    return income, expense, income - expense

def get_weekly_summary(user_id: int):
    week_ago = (datetime.utcnow().date() - timedelta(days=7)).isoformat()
    if DATABASE_URL:
        row = _execute_query("""
            SELECT 
                COALESCE(SUM(CASE WHEN type = 'income' THEN amount ELSE 0 END), 0) AS income,
                COALESCE(SUM(CASE WHEN type = 'expense' THEN amount ELSE 0 END), 0) AS expense
            FROM transactions
            WHERE user_id = %s AND created_at::date >= %s
        """, (user_id, week_ago))
    else:
        row = _execute_query("""
            SELECT 
                SUM(CASE WHEN type = 'income' THEN amount ELSE 0 END) AS income,
                SUM(CASE WHEN type = 'expense' THEN amount ELSE 0 END) AS expense
            FROM transactions
            WHERE user_id = ? AND DATE(created_at) >= ?
        """, (user_id, week_ago))
    
    income = (row["income"] if DATABASE_URL else row[0]) or 0
    expense = (row["expense"] if DATABASE_URL else row[1]) or 0
    return income, expense, income - expense

def get_monthly_summary(user_id: int):
    now = datetime.utcnow()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
    if DATABASE_URL:
        row = _execute_query("""
            SELECT 
                COALESCE(SUM(CASE WHEN type = 'income' THEN amount ELSE 0 END), 0) AS income,
                COALESCE(SUM(CASE WHEN type = 'expense' THEN amount ELSE 0 END), 0) AS expense
            FROM transactions
            WHERE user_id = %s AND created_at >= %s
        """, (user_id, month_start))
    else:
        row = _execute_query("""
            SELECT 
                SUM(CASE WHEN type = 'income' THEN amount ELSE 0 END) AS income,
                SUM(CASE WHEN type = 'expense' THEN amount ELSE 0 END) AS expense
            FROM transactions
            WHERE user_id = ? AND created_at >= ?
        """, (user_id, month_start))
    
    income = (row["income"] if DATABASE_URL else row[0]) or 0
    expense = (row["expense"] if DATABASE_URL else row[1]) or 0
    return income, expense, income - expense

def get_expense_categories_summary(user_id: int):
    conn = get_db_connection()
    cur = conn.cursor()
    if DATABASE_URL:
        cur.execute("""
            SELECT category, SUM(amount) AS total
            FROM transactions
            WHERE user_id = %s AND type = 'expense'
            GROUP BY category
            ORDER BY total DESC
        """, (user_id,))
        rows = cur.fetchall()
        result = [(row["category"], float(row["total"])) for row in rows]
    else:
        cur.execute("""
            SELECT category, SUM(amount) AS total
            FROM transactions
            WHERE user_id = ? AND type = 'expense'
            GROUP BY category
            ORDER BY total DESC
        """, (user_id,))
        rows = cur.fetchall()
        result = [(row[0], row[1]) for row in rows]
    conn.close()
    return result

# === EXCEL ГЕНЕРАЦИЯ ===

def generate_excel_report(user_id: int, filename: str):
    import os
    from openpyxl import Workbook
    from openpyxl.styles import Font

    conn = get_db_connection()
    cur = conn.cursor()
    
    # Операции
    if DATABASE_URL:
        cur.execute("""
            SELECT created_at, type, category, amount, description
            FROM transactions
            WHERE user_id = %s
            ORDER BY created_at
        """, (user_id,))
    else:
        cur.execute("""
            SELECT created_at, type, category, amount, description
            FROM transactions
            WHERE user_id = ?
            ORDER BY created_at
        """, (user_id,))
    operations = cur.fetchall()
    
    # Категории
    if DATABASE_URL:
        cur.execute("""
            SELECT category, SUM(amount) AS total
            FROM transactions
            WHERE user_id = %s AND type = 'expense'
            GROUP BY category
            ORDER BY total DESC
        """, (user_id,))
        category_summary = cur.fetchall()
        cat_list = [(row["category"], float(row["total"])) for row in category_summary]
    else:
        cur.execute("""
            SELECT category, SUM(amount) AS total
            FROM transactions
            WHERE user_id = ? AND type = 'expense'
            GROUP BY category
            ORDER BY total DESC
        """, (user_id,))
        category_summary = cur.fetchall()
        cat_list = [(row[0], row[1]) for row in category_summary]
    
    # Месяц
    now = datetime.utcnow()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
    if DATABASE_URL:
        cur.execute("""
            SELECT 
                COALESCE(SUM(CASE WHEN type = 'income' THEN amount ELSE 0 END), 0) AS income,
                COALESCE(SUM(CASE WHEN type = 'expense' THEN amount ELSE 0 END), 0) AS expense
            FROM transactions
            WHERE user_id = %s AND created_at >= %s
        """, (user_id, month_start))
        monthly = cur.fetchone()
        m_income = float(monthly["income"])
        m_expense = float(monthly["expense"])
    else:
        cur.execute("""
            SELECT 
                SUM(CASE WHEN type = 'income' THEN amount ELSE 0 END) AS income,
                SUM(CASE WHEN type = 'expense' THEN amount ELSE 0 END) AS expense
            FROM transactions
            WHERE user_id = ? AND created_at >= ?
        """, (user_id, month_start))
        monthly = cur.fetchone()
        m_income = monthly[0] or 0
        m_expense = monthly[1] or 0

    conn.close()

    # Создаём Excel
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Операции"
    ws1.append(["Дата и время", "Тип", "Категория", "Сумма", "Комментарий"])
    for row in operations:
        dt = str(row[0] if DATABASE_URL else row["created_at"]).split(".")[0]
        ws1.append([dt, row[1] if DATABASE_URL else row["type"], row[2] if DATABASE_URL else row["category"], float(row[3] if DATABASE_URL else row["amount"]), row[4] if DATABASE_URL else row["description"]])
    
    ws2 = wb.create_sheet("Сводка")
    ws2.append(["Категория", "Сумма"])
    for cat, total in cat_list:
        ws2.append([cat, total])
    
    ws2.append([])
    ws2.append(["Итого за месяц", ""])
    ws2.append(["Доход", m_income])
    ws2.append(["Расход", m_expense])
    ws2.append(["Прибыль", m_income - m_expense])
    
    for ws in [ws1, ws2]:
        for cell in ws[1]:
            cell.font = Font(bold=True)
    
    wb.save(filename)
